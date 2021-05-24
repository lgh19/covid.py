
#example of how to run code:
# a = CovidData('DOH_PASDA.csv')
# the PASDA file is passed into the CovidData class, which then scrapes the DOH LTFC data set and FPP data set from the web
#the output of the code is one csv file with the relevant columns of these 3 files  combined

from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import pandas as pd
from fuzzywuzzy import fuzz
import re


class CovidData:
    pasdaData = ""
    dohfile = ""
    fppfile = ''
    global ltcfdf
    global dohdf
    global fppdf
    global dohColumnName
    global doh_last_updated
    global fpp_last_updated

    def __init__(self, pasdaData):
        self.ltcfdf = pd.DataFrame
        self.ltcfdf = pd.read_csv(pasdaData)
        from bs4 import BeautifulSoup
        import requests

        r = requests.get("https://www.health.pa.gov/topics/disease/coronavirus/Pages/LTCF-Data.aspx")
        soup = BeautifulSoup(r.text, 'html.parser')

        doc_urls = []
        doh_dates = []
        for link in soup.find_all('a'):
            url = link.get('href', 'No Link Found')
            if re.search('xlsx', url) is not None:
                doh_dates.append(link.next_sibling)
                doc_urls.append(url)
        # print doc_urls
        # print doh_dates
        assert len(doc_urls) == 4  # Verify that the web site only lists four XSLX files.
        self.doh_last_updated = doh_dates[0]
        fullurl = "https://www.health.pa.gov" + doc_urls[0]
        print(fullurl)
        r = requests.get(fullurl)

        open('DOH_Data.xlsx', 'wb').write(r.content)

        r = requests.get("https://data.pa.gov/api/views/iwiy-rwzp/rows.csv?accessType=DOWNLOAD&api_foundry=true")
        open('FPP_Data.csv', 'wb').write(r.content)
        
    #    fullSite = requests.get("https://data.pa.gov/Health/COVID-19-Federal-Pharmacy-Partners-Long-Term-Care-/iwiy-rwzp")
    #    soup = BeautifulSoup(fullSite.text, 'html.parser')
     
        
      #  wb = load_workbook('FPP_Data.xlsx')
        wb2 = load_workbook('DOH_Data.xlsx')
     #   print(wb.sheetnames)
     #   ws = wb[wb.sheetnames[0]]

        ws2 = wb2['Sheet1']

     #   df = pd.DataFrame(ws.values)

     #   df.rename(
      #      columns={0: u"Facility Name", 1: u'Address', 2: u"City", 3: u"County", 4: u"Zip Code", 5: u'Facility Type',
     #                6: u"clinicdt1",
     #                7: u"clinicdt2", 8: u"clinicdt3", 9: u"clinicdt4", 10: u"clinicdt5", 11: u"clinicdt6",
      #               12: u"clinicdt7", 13: u"clinicdt8", 14: u"clinicdt9", 15: u"clinicdt10"}, inplace=True)

        df2 = pd.DataFrame(ws2.values)

        df2.rename(columns={0: u"FACID", 1: u"NAME", 2: u"CITY", 3: u"COUNTY", 4: u"ALL_BEDS", 5: u"CURRENT_CENSUS",
                            6: u"Resident Cases to Display",
                            7: u"Resident Deaths to Display", 8: u"Staff Cases to Display"}, inplace=True)

     #   df = df.iloc[1:]
        df2 = df2.iloc[1:]

      #  df.to_csv("FPP_CSV1.csv", encoding='utf-8')

        df2.to_csv("DOH_CSV1.csv", encoding='utf-8')
        self.fppdf = pd.read_csv('FPP_Data.csv')
        self.dohdf = pd.read_csv('DOH_CSV1.csv')
        self.addDOHData()

    def addDOHData(self):
        print
        print("Adding DOH Covid Data")
        print

        def isNaN(string):
            return string != string

        self.dohdf.rename(columns={'Resident Cases to Display': 'Resident_Cases_to_Display'}, inplace=True)
        self.dohdf.rename(columns={'Resident Deaths to Display': 'Resident_Deaths_to_Display'}, inplace=True)
        self.dohdf.rename(columns={'Staff Cases to Display': 'Staff_Cases_to_Display'}, inplace=True)
        self.ltcfdf['ALL_BEDS'] = ''
        self.ltcfdf['CURRENT_CENSUS'] = ''
        self.ltcfdf['Resident_Cases_to_Display'] = ''
        self.ltcfdf['Resident_Deaths_to_Display'] = ''
        self.ltcfdf['Staff_Cases_to_Display'] = ''

        for index, row in self.ltcfdf.iterrows():
            tempfacid = row.FACILITY_I
            tempbeds = ''
            tempcensus = ''
            tempres = ''
            tempresdeath = ''
            tempstaff = ''
            for index, row in self.dohdf.iterrows():
                if row.FACID != tempfacid:
                    pass
                else:
                    # print "match"
                    tempbeds = row.ALL_BEDS
                    tempcensus = row.CURRENT_CENSUS
                    tempres = row.Resident_Cases_to_Display
                    tempresdeath = row.Resident_Deaths_to_Display
                    tempstaff = row.Staff_Cases_to_Display

                    self.ltcfdf.loc[self.ltcfdf['FACILITY_I'] == tempfacid, 'ALL_BEDS'] = tempbeds
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_I'] == tempfacid, 'CURRENT_CENSUS'] = tempcensus
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_I'] == tempfacid, 'Resident_Cases_to_Display'] = tempres
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_I'] == tempfacid, 'Resident_Deaths_to_Display'] = tempresdeath
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_I'] == tempfacid, 'Staff_Cases_to_Display'] = tempstaff

        self.ltcfdf['DOH_Data_Last_Updated'] = self.doh_last_updated
        self.ltcfdf = self.ltcfdf
        print("Done Adding DOH Covid Data")
        print
        self.addFPPData()

    def addFPPData(self):
        print("Adding FPP Data")
        print

        print
        for index, row in self.fppdf.iterrows():
            tempStr = (self.fppdf.at[index, 'Long Term Care Facility Name'])
            tempStr = str(tempStr)
            tempfacility = self.fppdf.at[index, 'Long Term Care Facility Name']
            tempaddress = self.fppdf.at[index, 'Street Address']
            tempPharm = self.fppdf.at[index, 'Federal Pharmacy Partner']
            date1 = self.fppdf.at[index, '1st Clinic Date']
            date2 = self.fppdf.at[index, '2nd Clinic Date']
            date3 = self.fppdf.at[index, '3rd Clinic Date']
            date4 = self.fppdf.at[index, '4th Clinic Date']
            date5 = self.fppdf.at[index, '5th Clinic Date']
            date6 = self.fppdf.at[index, '6th Clinic Date']
            date7 = self.fppdf.at[index, '7th Clinic Date']
            date8 = self.fppdf.at[index, '8th Clinic Date']
            date9 = self.fppdf.at[index, '9th Clinic Date']
            date10 = self.fppdf.at[index, '10th Clinic Date']
            tempTotal = self.fppdf.at[index, 'Total Doses Administered']
            tempFirstDose = self.fppdf.at[index,'First Doses Administered']
            tempSecondDose = self.fppdf.at[index,'Second Doses Adminstered']
            tempTotalRes = self.fppdf.at[index, 'Total Resident Doses Administered']
            tempTotalStaff = self.fppdf.at[index, 'Total Staff Doses Administered']

            for index, row in self.ltcfdf.iterrows():
                Str1 = (self.ltcfdf.at[index, 'FACILITY_N'])
                Str1 = str(Str1)
                ratio = fuzz.ratio(Str1.lower(), tempStr.lower())
                if (ratio >= 90):
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Facility Name FPP'] = tempfacility
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Federal Pharmacy Partner'] = tempPharm
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'First Clinic Date '] = date1
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Second Clinic Date'] = date2
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Third Clinic Date'] = date3
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Fourth Clinic Date'] = date4
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Fifth Clinic Date'] = date5
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Sixth Clinic Date'] = date6
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Seventh Clinic Date'] = date7
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Eighth Clinic Date'] = date8
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Ninth Clinic Date'] = date9
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Tenth Clinic Date'] = date10
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Total Doses Administered'] = tempTotal
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'First Doses Administered'] = tempFirstDose
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Second Doses Adminstered'] = tempSecondDose
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Total Resident Doses Administered'] = tempTotalRes
                    self.ltcfdf.loc[self.ltcfdf['FACILITY_N'] == Str1, 'Total Staff Doses Administered'] = tempTotalStaff
                    break

      #  self.ltcfdf['FPP_Data_Last_Updated'] = self.fpp_last_updated
        self.ltcfdf = self.ltcfdf
        print("Done Adding FPP Data")
        self.writeToFile()

    def writeToFile(self):
        print
        b = "COVID19_Vaccine_Data_LTCF.csv"
        b = str(b)
        self.ltcfdf.to_csv(b, encoding='utf-8')
        print("output file: "),
        print(b)
        return b

a = CovidData("Pasda.csv")
